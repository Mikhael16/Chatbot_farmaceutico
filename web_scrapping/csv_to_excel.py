#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script para convertir CSV de Inkafarma a Excel con formato y documentación.

Uso:
  python csv_to_excel.py <archivo_csv> <archivo_excel>
  python csv_to_excel.py inkafarma_final_test.csv inkafarma_productos.xlsx
"""

import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_excel_from_csv(csv_path, excel_path):
    """Convierte CSV a Excel con formato profesional."""
    
    # Leer CSV
    print(f"Leyendo CSV: {csv_path}")
    df = pd.read_csv(csv_path, encoding='utf-8')
    
    # Crear Excel con múltiples hojas
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Hoja 1: Datos de productos
        df.to_excel(writer, sheet_name='Productos', index=False)
        
        # Hoja 2: Estadísticas
        stats_data = {
            'Métrica': [
                'Total de productos',
                'Productos con precio',
                'Productos disponibles',
                'Productos agotados',
                'Categorías únicas',
                'Productos Venta Libre',
                'Productos con Receta',
                'Precio promedio',
                'Precio mínimo',
                'Precio máximo'
            ],
            'Valor': [
                len(df),
                len(df[df['precio_promocion'].notna() & (df['precio_promocion'] != '')]),
                len(df[df['stock_disponible'] == 'Disponible']),
                len(df[df['stock_disponible'] == 'Agotado']),
                df['categoria'].nunique(),
                len(df[df['condicion_venta'] == 'Venta Libre']),
                len(df[df['condicion_venta'].str.contains('Receta', na=False)]),
                df['precio_promocion'].replace('', pd.NA).astype(float, errors='ignore').mean(),
                df['precio_promocion'].replace('', pd.NA).astype(float, errors='ignore').min(),
                df['precio_promocion'].replace('', pd.NA).astype(float, errors='ignore').max()
            ]
        }
        pd.DataFrame(stats_data).to_excel(writer, sheet_name='Estadísticas', index=False)
        
        # Hoja 3: Documentación
        doc_data = {
            'Campo': [
                'sku',
                'nombre_comercial',
                'dci',
                'forma',
                'concentracion',
                'presentacion',
                'condicion_venta',
                'registro_sanitario',
                'precio_lista',
                'precio_promocion',
                'categoria',
                'subcategoria',
                'stock_disponible',
                'url_producto',
                'fecha_extraccion'
            ],
            'Descripción': [
                'Código SKU único del producto',
                'Nombre comercial del medicamento/producto',
                'Denominación Común Internacional (principio activo)',
                'Forma farmacéutica (tableta, jarabe, cápsula, etc.)',
                'Concentración del principio activo',
                'Presentación comercial (caja x20, frasco 100ml, etc.)',
                'Condición de venta (Venta Libre, Receta Médica)',
                'Registro sanitario DIGEMID',
                'Precio de lista/regular (antes de descuento)',
                'Precio de venta/promoción actual',
                'Categoría principal del producto',
                'Subcategoría del producto',
                'Estado de disponibilidad (Disponible/Agotado)',
                'URL de la página del producto',
                'Fecha y hora de extracción ISO 8601'
            ],
            'Tipo': [
                'Texto',
                'Texto',
                'Texto',
                'Texto',
                'Texto',
                'Texto',
                'Texto',
                'Texto',
                'Numérico',
                'Numérico',
                'Texto',
                'Texto',
                'Texto',
                'URL',
                'DateTime'
            ],
            'Obligatorio': [
                'Sí',
                'Sí',
                'No',
                'No',
                'No',
                'No',
                'Sí',
                'No',
                'No',
                'Sí',
                'Sí',
                'No',
                'Sí',
                'Sí',
                'Sí'
            ]
        }
        pd.DataFrame(doc_data).to_excel(writer, sheet_name='Documentación', index=False)
        
        # Hoja 4: Metadatos
        metadata = {
            'Propiedad': [
                'Proyecto',
                'Fuente de datos',
                'Fecha de extracción',
                'Herramienta',
                'Versión',
                'Total de registros',
                'Categorías',
                'Responsable',
                'Propósito',
                'Nota técnica'
            ],
            'Valor': [
                'Chatbot Farmacéutico - Inkafarma',
                'https://inkafarma.pe',
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Web Scraper con Playwright',
                '1.0',
                str(len(df)),
                ', '.join(df['categoria'].unique()),
                'Equipo Digital Innovation',
                'Alimentar base de datos del chatbot farmacéutico',
                'Datos extraídos con renderizado JavaScript (Angular)'
            ]
        }
        pd.DataFrame(metadata).to_excel(writer, sheet_name='Metadatos', index=False)
    
    # Aplicar formato
    print(f"Aplicando formato al archivo: {excel_path}")
    wb = load_workbook(excel_path)
    
    # Formato para hoja de Productos
    ws_productos = wb['Productos']
    format_sheet(ws_productos, header_fill='1F4E78', header_font_color='FFFFFF')
    
    # Formato para hoja de Estadísticas
    ws_stats = wb['Estadísticas']
    format_sheet(ws_stats, header_fill='4472C4', header_font_color='FFFFFF')
    
    # Formato para hoja de Documentación
    ws_doc = wb['Documentación']
    format_sheet(ws_doc, header_fill='70AD47', header_font_color='FFFFFF')
    
    # Formato para hoja de Metadatos
    ws_meta = wb['Metadatos']
    format_sheet(ws_meta, header_fill='FFC000', header_font_color='000000')
    
    # Guardar
    wb.save(excel_path)
    print(f"✓ Excel creado exitosamente: {excel_path}")
    print(f"  - {len(df)} productos")
    print(f"  - 4 hojas: Productos, Estadísticas, Documentación, Metadatos")


def format_sheet(ws, header_fill='4472C4', header_font_color='FFFFFF'):
    """Aplica formato profesional a una hoja de Excel."""
    
    # Estilo de encabezados
    header_fill_style = PatternFill(start_color=header_fill, end_color=header_fill, fill_type='solid')
    header_font_style = Font(bold=True, color=header_font_color, size=11)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aplicar a encabezados
    for cell in ws[1]:
        cell.fill = header_fill_style
        cell.font = header_font_style
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_style
    
    # Auto-ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    cell.border = border_style
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        # Limitar ancho máximo
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Congelar primera fila
    ws.freeze_panes = 'A2'


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Uso: python csv_to_excel.py <archivo_csv> <archivo_excel>")
        print("Ejemplo: python csv_to_excel.py inkafarma_final_test.csv inkafarma_productos.xlsx")
        sys.exit(1)
    
    csv_file = sys.argv[1]
    excel_file = sys.argv[2]
    
    create_excel_from_csv(csv_file, excel_file)
